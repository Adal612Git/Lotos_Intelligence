import argparse
import csv
import json
import threading
import logging
import os
import re
import time
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ==========================
# CONFIG (NO KEYS REQUIRED)
# ==========================
MUNICIPIOS = [
    "Guadalajara",
    "Zapopan",
]

RUBROS = [
    "Restaurantes",
    "Cafeterias",
    "Talleres Mecanicos",
    "Dentistas",
    "Veterinarias",
    "Gimnasios",
    "Salones de belleza",
    "Farmacias",
]

OUTPUT_XLSX = "Reporte_Lotos.xlsx"
PARTIAL_CSV = "Reporte_Lotos_Parcial.csv"
LOG_FILE = "lotos_execution.log"

OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.nchc.org.tw/api/interpreter",
]

MAX_TOTAL_RECORDS = 5000
PARTIAL_FLUSH_EVERY = 500
SLEEP_BETWEEN_CALLS = 1.0
HTTP_TIMEOUT_SEC = 20
SPINNER_INTERVAL = 0.3
HTTP_TIMEOUT_SEC = 60

SUMMARY_SHEET = "Resumen"
CHARTS_SHEET = "Graficas"

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)


def clean_text(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def normalize_phone(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"[^0-9+]", "", text)


def has_website(url: str) -> str:
    return "Si" if url else "No"


def osm_tag_map() -> Dict[str, List[Tuple[str, str]]]:
    return {
        "Restaurantes": [("amenity", "restaurant")],
        "Cafeterias": [("amenity", "cafe")],
        "Notarias": [("office", "notary")],
        "Talleres Mecanicos": [("amenity", "car_repair"), ("shop", "car_repair")],
        "Spas": [("leisure", "spa"), ("amenity", "spa")],
        "Dentistas": [("amenity", "dentist")],
        "Clinicas": [("amenity", "clinic"), ("healthcare", "clinic")],
        "Veterinarias": [("amenity", "veterinary")],
        "Gimnasios": [("leisure", "fitness_centre"), ("amenity", "gym")],
        "Salones de belleza": [("shop", "beauty"), ("shop", "hairdresser")],
        "Inmobiliarias": [("office", "real_estate_agent")],
        "Agencias de viajes": [("shop", "travel_agency")],
        "Escuelas privadas": [("amenity", "school"), ("amenity", "college")],
        "Consultorios psicologicos": [("healthcare", "psychotherapist"), ("office", "psychotherapist")],
        "Abogados": [("office", "lawyer")],
        "Laboratorios clinicos": [("healthcare", "laboratory")],
        "Contadores": [("office", "accountant")],
        "Funerarias": [("amenity", "funeral_home")],
        "Papelerias": [("shop", "stationery")],
        "Ferreterias": [("shop", "hardware")],
        "Floristerias": [("shop", "florist")],
        "Panaderias": [("shop", "bakery")],
        "Farmacias": [("amenity", "pharmacy")],
        "Opticas": [("shop", "optician")],
        "Tiendas de electronica": [("shop", "electronics")],
    }


def overpass_query(municipio: str, tag_filters: List[Tuple[str, str]]) -> str:
    parts = []
    for k, v in tag_filters:
        parts.append(f'nwr["{k}"="{v}"](area.a);')
    filters = "\n  ".join(parts)
    return (
        '[out:json][timeout:50];\n'
        f'area["name"="{municipio}"]["admin_level"="8"]["boundary"="administrative"]->.a;\n'
        "(\n"
        f"  {filters}\n"
        ");\n"
        "out tags center;"
    )


def http_post_json(url: str, query: str) -> Dict:
    try:
        import urllib.request

        data = query.encode("utf-8")
        req = urllib.request.Request(url, data=data, method="POST")
        req.add_header("Content-Type", "application/x-www-form-urlencoded; charset=UTF-8")
        with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT_SEC) as resp:
            text = resp.read().decode("utf-8", errors="ignore")
            return json.loads(text) if text else {}
    except Exception as exc:
        logging.error(f"HTTP error: {exc}")
        return {}


def call_overpass(query: str) -> Dict:
    for endpoint in OVERPASS_ENDPOINTS:
        stop_event = threading.Event()
        spinner_thread = threading.Thread(
            target=spinner,
            args=(f"Esperando Overpass ({endpoint.split('//')[1]})", stop_event),
            daemon=True,
        )
        spinner_thread.start()
        data = http_post_json(endpoint, query)
        stop_event.set()
        spinner_thread.join(timeout=1)
        if data:
            return data
        time.sleep(1.0)
    return {}


def spinner(label: str, stop_event: threading.Event) -> None:
    frames = ["|", "/", "-", "\\"]
    i = 0
    while not stop_event.is_set():
        print(f"\r{label} {frames[i % len(frames)]}", end="", flush=True)
        time.sleep(SPINNER_INTERVAL)
        i += 1
    print("\r" + " " * (len(label) + 4) + "\r", end="", flush=True)


def tags_to_row(rubro: str, municipio: str, el: Dict) -> Dict[str, str]:
    tags = el.get("tags", {}) or {}
    name = clean_text(tags.get("name", ""))
    phone = clean_text(tags.get("phone", "")) or clean_text(tags.get("contact:phone", ""))
    web = clean_text(tags.get("website", "")) or clean_text(tags.get("contact:website", ""))
    email = clean_text(tags.get("email", "")) or clean_text(tags.get("contact:email", ""))
    street = clean_text(tags.get("addr:street", ""))
    housenumber = clean_text(tags.get("addr:housenumber", ""))
    suburb = clean_text(tags.get("addr:suburb", ""))
    postcode = clean_text(tags.get("addr:postcode", ""))
    lat = str(el.get("lat") or (el.get("center") or {}).get("lat") or "")
    lon = str(el.get("lon") or (el.get("center") or {}).get("lon") or "")

    return {
        "nombre": name,
        "telefono": normalize_phone(phone),
        "sitio_web": web,
        "correo": email,
        "actividad": rubro,
        "direccion": f"{street} {housenumber}".strip(),
        "colonia": suburb,
        "cp": postcode,
        "municipio": municipio,
        "entidad": "Jalisco",
        "latitud": lat,
        "longitud": lon,
        "tiene_web": has_website(web),
        "fuente": "OpenStreetMap",
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def dedupe_key(row: Dict[str, str]) -> str:
    return "|".join(
        [
            clean_text(row.get("nombre", "")).lower(),
            row.get("latitud", ""),
            row.get("longitud", ""),
        ]
    )


def flush_partial(rows: List[Dict[str, str]]) -> None:
    if not rows:
        return
    file_exists = os.path.exists(PARTIAL_CSV)
    with open(PARTIAL_CSV, "a", newline="", encoding="utf-8") as f:
        writer = json_csv_writer(rows, f, file_exists)
        writer.writerows(rows)


def json_csv_writer(rows: List[Dict[str, str]], f, file_exists: bool):
    keys = list(rows[0].keys())
    writer = csv.DictWriter(f, fieldnames=keys)
    if not file_exists:
        writer.writeheader()
    return writer


def build_report(df: pd.DataFrame) -> None:
    if df.empty:
        return
    for col in ["tiene_web", "telefono", "actividad", "municipio"]:
        if col not in df.columns:
            df[col] = ""
    df.to_excel(OUTPUT_XLSX, index=False, sheet_name="Data")
    wb = load_workbook(OUTPUT_XLSX)
    data_ws = wb["Data"]

    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    if CHARTS_SHEET in wb.sheetnames:
        del wb[CHARTS_SHEET]

    summary_ws = wb.create_sheet(SUMMARY_SHEET)
    charts_ws = wb.create_sheet(CHARTS_SHEET)

    total = len(df)
    df["tiene_web"] = df["tiene_web"].fillna("").astype(str)
    df["telefono"] = df["telefono"].fillna("").astype(str)
    pct_web = round((df["tiene_web"].str.upper() == "SI").mean() * 100, 2)
    pct_phone = round((df["telefono"].str.len() > 0).mean() * 100, 2)

    summary_ws["A1"] = "Resumen Ejecutivo"
    summary_ws["A3"] = "Total registros"
    summary_ws["B3"] = total
    summary_ws["A4"] = "Porcentaje con web"
    summary_ws["B4"] = pct_web
    summary_ws["A5"] = "Porcentaje con telefono"
    summary_ws["B5"] = pct_phone

    rubro_counts = df.groupby("actividad").size().sort_values(ascending=False)
    municipio_counts = df.groupby("municipio").size().sort_values(ascending=False)

    row = 7
    summary_ws[f"A{row}"] = "Rubros (conteo)"
    for i, (k, v) in enumerate(rubro_counts.items(), start=1):
        summary_ws[f"A{row + i}"] = k
        summary_ws[f"B{row + i}"] = int(v)

    row = row + len(rubro_counts) + 2
    summary_ws[f"A{row}"] = "Municipios (conteo)"
    for i, (k, v) in enumerate(municipio_counts.items(), start=1):
        summary_ws[f"A{row + i}"] = k
        summary_ws[f"B{row + i}"] = int(v)

    rubro_start = 8
    rubro_end = 7 + len(rubro_counts)
    if rubro_end >= rubro_start:
        bar = BarChart()
        bar.title = "Rubros (conteo)"
        data = Reference(summary_ws, min_col=2, min_row=rubro_start, max_row=rubro_end)
        cats = Reference(summary_ws, min_col=1, min_row=rubro_start, max_row=rubro_end)
        bar.add_data(data, titles_from_data=False)
        bar.set_categories(cats)
        bar.height = 10
        bar.width = 20
        charts_ws.add_chart(bar, "A1")

    mun_start = rubro_end + 3
    mun_end = mun_start + len(municipio_counts) - 1
    if mun_end >= mun_start:
        pie = PieChart()
        pie.title = "Municipios"
        data = Reference(summary_ws, min_col=2, min_row=mun_start, max_row=mun_end)
        cats = Reference(summary_ws, min_col=1, min_row=mun_start, max_row=mun_end)
        pie.add_data(data, titles_from_data=False)
        pie.set_categories(cats)
        pie.height = 10
        pie.width = 14
        charts_ws.add_chart(pie, "A20")

    for ws in [data_ws, summary_ws]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(OUTPUT_XLSX)


def main() -> None:
    tag_map = osm_tag_map()
    rows: List[Dict[str, str]] = []
    seen = set()
    total = 0
    completed_calls = 0
    total_calls = len(MUNICIPIOS) * len(RUBROS)

    print(f"Iniciando: {len(MUNICIPIOS)} municipios x {len(RUBROS)} rubros")
    if os.path.exists(PARTIAL_CSV):
        try:
            df_old = pd.read_csv(PARTIAL_CSV, nrows=1)
            for col in ["tiene_web", "telefono", "actividad", "municipio"]:
                if col not in df_old.columns:
                    os.remove(PARTIAL_CSV)
                    print("CSV parcial viejo incompatible. Se reinicia.")
                    break
        except Exception:
            os.remove(PARTIAL_CSV)
            print("CSV parcial dañado. Se reinicia.")
    def progress_bar(done: int, totalc: int, width: int = 30) -> str:
        if totalc <= 0:
            return "[------------------------------]"
        filled = int((done / totalc) * width)
        return "[" + "#" * filled + "-" * (width - filled) + "]"

    for municipio in MUNICIPIOS:
        for rubro in RUBROS:
            tag_filters = tag_map.get(rubro, [])
            if not tag_filters:
                continue

            print(f"Consultando: {rubro} en {municipio}")
            start_t = time.time()
            query = overpass_query(municipio, tag_filters)
            data = call_overpass(query)
            elements = data.get("elements", []) if data else []
            completed_calls += 1

            for el in elements:
                row = tags_to_row(rubro, municipio, el)
                key = dedupe_key(row)
                if key and key in seen:
                    continue
                if key:
                    seen.add(key)
                rows.append(row)
                total += 1

                if total >= MAX_TOTAL_RECORDS:
                    break

            if len(rows) >= PARTIAL_FLUSH_EVERY:
                flush_partial(rows)
                rows = []

            if total >= MAX_TOTAL_RECORDS:
                break

            pct = round((completed_calls / total_calls) * 100, 2)
            elapsed = round(time.time() - start_t, 1)
            bar = progress_bar(completed_calls, total_calls, 30)
            print(
                f"{bar} {completed_calls}/{total_calls} ({pct}%) | "
                f"registros: {total} | {elapsed}s"
            )

            time.sleep(SLEEP_BETWEEN_CALLS)

        if total >= MAX_TOTAL_RECORDS:
            break

    if rows:
        flush_partial(rows)

    if os.path.exists(PARTIAL_CSV):
        df = pd.read_csv(PARTIAL_CSV)
        build_report(df)
        print(f"Reporte generado: {OUTPUT_XLSX}")
    else:
        print("No se encontraron registros.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--selftest", action="store_true", help="Prueba completa rapida")
    parser.add_argument("--smoke", action="store_true", help="Prueba minima (1 municipio, 1 rubro)")
    args = parser.parse_args()

    if args.selftest:
        MUNICIPIOS[:] = ["Guadalajara"]
        RUBROS[:] = ["Restaurantes"]
        MAX_TOTAL_RECORDS = 50
        HTTP_TIMEOUT_SEC = 15
        SLEEP_BETWEEN_CALLS = 0.5
        print("Selftest: Guadalajara + Restaurantes")
        main()
    elif args.smoke:
        MUNICIPIOS[:] = ["Guadalajara"]
        RUBROS[:] = ["Cafeterias"]
        MAX_TOTAL_RECORDS = 20
        HTTP_TIMEOUT_SEC = 10
        SLEEP_BETWEEN_CALLS = 0.3
        print("Smoke test: Guadalajara + Cafeterias")
        main()
    else:
        main()
